import os
import openpyxl
import pandas as pd
import math
import argparse
import socket
import glob

# DONE
#  - ホスト名の登録、取り消し処理の実装
# TODO

task = 'distractor'

class node_comb:

    '''
    comb[0]...組み合わせ例, シート名
    comb[1]...シートのデータ数. index無し（まっさら）は-1, indexありは0
    '''

    def __init__(self, path):

        self.path = path
        self.path_excel = path + 'result_' + task + '.xlsx'
        self.log_command = 'log_command.txt'

        self.comb = [
            ['RRRRRR', 'LRRRRR', 'RLRRRR', 'LLRRRR', 'RRLLLL', 'LRLLLL', 'RLLLLL', 'LLLLLL'],
            [-1, -1, -1, -1, -1, -1, -1, -1]
        ]

        if os.path.isfile(self.path_excel) == False:
            excel = openpyxl.Workbook()
            excel.save(self.path_excel)
            excel.close()
        self.processor_init()
        self.load_comb()


    # データシートの初期化 (replaceで、存在するシートもリセット)
    def processor_init(self):

        sheet_name = 'process'
        excel = openpyxl.load_workbook(self.path_excel)
        sheets = excel.sheetnames
        excel.close()
        if 'process' not in sheets:
            host = ['-' for i in range(len(self.comb[0]))]
            df = pd.DataFrame(index=self.comb[0])
            df['HOST'] = '-'
            with pd.ExcelWriter(self.path_excel, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name)
            excel = openpyxl.load_workbook(self.path_excel)
            del excel['Sheet']
            excel.save(self.path_excel)
            excel.close()
        else:
            df = pd.read_excel(self.path_excel, sheet_name=sheet_name, index_col=0)
            if(len(df) == 0):
                df = pd.DataFrame(index=self.comb[0])
                df['HOST'] = '-'
                with pd.ExcelWriter(self.path_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name)


    # combのデータ数が読み込みされます
    def load_comb(self):

        excel = openpyxl.load_workbook(self.path_excel)
        sheets = excel.sheetnames
        excel.close()

        for i in range(len(self.comb[0])):  # 規定の組み合わせを検索
            found_sheet = False
            for sheet in sheets:
                if sheet == self.comb[0][i]:  # シート名が規定の組み合わせと一致
                    found_sheet = True
                    self.comb[1][i] = self.check_train_num(i)
                    if self.comb[1][i] == -1:   # シートが存在するだけの場合
                        self.init_sheet(i, replace=True)
                    break
            if found_sheet == False:
                self.init_sheet(i)


    # データ数を取得 (シートなし(-1)、シートあり/データなし(0))
    def check_train_num(self, sheet_index):

        df = pd.read_excel(self.path_excel, self.comb[0][sheet_index], index_col=0)
        if len(df) != 0:
            return len(df.columns)
        return -1


    def init_sheet(self, sheet_index, replace=False):

        df = pd.DataFrame(index=[
            'em', 'f1', 'prec', 'recall',
            'sp_em', 'sp_f1', 'sp_prec', 'sp_recall',
            'joint_em', 'joint_f1', 'joint_prec', 'joint_recall',
            'last_eval', 'last_dev_loss', 'para_limit',
            'batch_size', 'init_lr', 'keep_prob', 'sp_lambda',
            'HOST_NAME', 'DATE'])
        if replace is True:
            with pd.ExcelWriter(self.path_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=self.comb[0][sheet_index])
        elif replace is False:
            with pd.ExcelWriter(self.path_excel, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=self.comb[0][sheet_index])
        self.comb[1][sheet_index] = 0


    def start_process(self, prep=False, tr=True, pred=True, eva=True, process_num=1, para_limit=2250, batch_size=24, init_lr=0.1, keep_prob=1.0, sp_lambda=1.0, task='distractor', date=None, train_target=None):

        self.para_limit = para_limit
        self.batch_size = batch_size
        self.init_lr = init_lr
        self.keep_prob = keep_prob
        self.sp_lambda = sp_lambda
        self.task = task
        self.date = date
        self.manage_host(mode='reset')
        self.manage_host(mode='show')
        os.environ['CUDA_VISIBLE_DEVICES'] = '0'

        for i in range(process_num):
            if train_target == None:
                self.train_target = self.select_comb()
            else:
                self.train_target = train_target
            self.manage_host(mode='regist')
            self.manage_host(mode='show')

            if prep is True:
                self.prepro()
            if tr is True:
                self.train()
            if pred is True:
                self.predict()
            if eva is True:

                self.output(self.evaluate())
            self.manage_host(mode='reset')
            self.manage_host(mode='show')


    def prepro(self, para_limit=None):

        print('\n[prepro]')
        print('='*30)
        glove_word_file = self.path + 'glove.840B.300d.txt'
        if para_limit == None:
           para_limit = self.para_limit
        command = 'python {}main.py --mode prepro --data_file {}hotpot_train_v1.1.json --para_limit {} --data_split train --glove_word_file {}'.format(self.path, self.path, para_limit, glove_word_file)
        self.write_log(self.log_command, command)
        os.system(command)
        command = 'python {}main.py --mode prepro --data_file {}hotpot_dev_distractor_v1.json --para_limit {} --data_split dev --glove_word_file {}'.format(self.path, self.path, para_limit, glove_word_file)
        self.write_log(self.log_command, command)
        os.system(command)
        command = 'python {}main.py --mode prepro --data_file {}hotpot_dev_fullwiki_v1.json --data_split dev --fullwiki --para_limit {} --glove_word_file {}'.format(self.path, self.path, para_limit, glove_word_file)
        self.write_log(self.log_command, command)
        os.system(command)
        print('[prepro] process has done.')


    def train(self, train_target=None, para_limit=None, batch_size=None, init_lr=None, keep_prob=None, sp_lambda=None):

        print('\n[train]')
        print('='*30)
        if train_target == None:
            train_target = self.train_target
        if para_limit == None:
           para_limit = self.para_limit
        if batch_size == None:
           batch_size = self.batch_size
        if init_lr == None:
           init_lr = self.init_lr
        if keep_prob == None:
           keep_prob = self.keep_prob
        if sp_lambda == None:
           sp_lambda = self.sp_lambda

        command = 'python {}main.py --mode train --para_limit {} --batch_size {} --init_lr {} --keep_prob {} --sp_lambda {} --train_target {}'.format(self.path, para_limit, batch_size, init_lr, keep_prob, sp_lambda, train_target)
        self.write_log(self.log_command, command)
        print('running command...\n  {}\n'.format(command))
        os.system(command)

        print('[train] process has done.')


    def predict(self, train_target=None, task=None, para_limit=None, batch_size=None, init_lr=None, keep_prob=None, sp_lambda=None, date=None):

        print('\n[predict]')
        print('='*30)
        if train_target == None:
            train_target = self.train_target
        if task == None:
            task = self.task
        if para_limit == None:
            para_limit = self.para_limit
        if batch_size == None:
            batch_size = self.batch_size
        if init_lr == None:
            init_lr = self.init_lr
        if keep_prob == None:
            keep_prob = self.keep_prob
        if sp_lambda == None:
            sp_lambda = self.sp_lambda
        if date == None:
            if self.date == None:
                path = self.path + 'HOTPOT-*'
                date = sorted(glob.glob(path), reverse=True)[-1][-15:]
                self.date = date
            else:
                date = self.date
        prediction_file = self.path + 'dev_' + task + '_pred.json'
        if (task == 'distractor') or (task == 'all'):
            command = 'python {}main.py --mode test --data_split dev --para_limit {} --batch_size {} --init_lr {} --keep_prob {} --sp_lambda {} --save HOTPOT-{} --train_target {} --prediction_file {}'.format(self.path, para_limit, batch_size, init_lr, keep_prob, sp_lambda, date, train_target, prediction_file)
            self.write_log(self.log_command, command)
            print('running command...\n  {}\n'.format(command))
            os.system(command)
        if (task == 'fullwiki') or (task == 'all'):
            command = 'python {}main.py --mode test --data_split dev --para_limit {} --batch_size {} --init_lr {} --keep_prob {} --sp_lambda {} --save HOTPOT-{} --train_target {} --prediction_file {}'.format(self.path, para_limit, batch_size, init_lr, keep_prob, sp_lambda, date, train_target, prediction_file)
            self.write_log(self.log_command, command)
            print('running command...\n  {}\n'.format(command))
            os.system(command)
        print('[predict] process has done.')


    def evaluate(self, task=None, para_limit=None, batch_size=None, init_lr=None, keep_prob=None, sp_lambda=None, date=None):

        print('\n[evaluate]')
        print('='*30)
        if task == None:
            task = self.task
        if para_limit == None:
           para_limit = self.para_limit
        if batch_size == None:
           batch_size = self.batch_size
        if init_lr == None:
           init_lr = self.init_lr
        if keep_prob == None:
           keep_prob = self.keep_prob
        if sp_lambda == None:
           sp_lambda = self.sp_lambda

        result_text = 'result_' + task + '.txt'
        command = 'python {}hotpot_evaluate_v1.py dev_{}_pred.json hotpot_dev_{}_v1.json > {}'.format(self.path, task, task, result_text)
        self.write_log(self.log_command, command)
        os.system(command)

        if date == None:
            if self.date == None:
                date = self.path + 'HOTPOT-*'
                date = sorted(glob.glob(date), reverse=True)[-1][-15:]
                self.date = date
            else:
                date = self.date

        print('getting last_eval, last_dev_loss ...')
        trainlog_text = self.path + 'HOTPOT-' + date + '\\log.txt'
        with open(trainlog_text) as f:
            train_lines = f.readlines()
        train_lines = [train_line.strip() for train_line in train_lines]
        train_line = [train_line for train_line in train_lines if 'eval' in train_line][-1].strip('|')
        data = train_line.split('|')
        for datos in data:
            datos = datos.split(':')
            index = datos[0].strip(' ')
            value = datos[1]
            if index == 'eval':
                value = value.strip(' ')[:2]
                last_eval = value
                print('last_eval: {}'.format(last_eval))
                last_eval = int(last_eval)
            elif index == 'dev loss':
                last_dev_loss = value
                print('last_dev_loss: {}'.format(last_dev_loss))
                last_dev_loss = float(last_dev_loss)
            else:
                print('{}: {}'.format(index, value))

        print('getting scores ...')
        with open(result_text) as f:
            lines = f.readlines()
            lines = [line.strip() for line in lines]
            lines = [line for line in lines if '{' in line]
            line = lines[0]
            data = line.split(',')
            index, value = [], []
            print('reshaping scores ...')
            for datos in data:
                datos = datos.split(':')
                index.append(datos[0].strip('{ \''))
                value.append(float(datos[1].strip('} ')))
            index.append('last_eval')
            value.append(last_eval)
            index.append('last_dev_loss')
            value.append(last_dev_loss)
            index.append('para_limit')
            value.append(para_limit)
            index.append('batch_size')
            value.append(batch_size)
            index.append('init_lr')
            value.append(init_lr)
            index.append('keep_prob')
            value.append(keep_prob)
            index.append('sp_lambda')
            value.append(sp_lambda)
            index.append('HOST_NAME')
            value.append(socket.gethostname())
            index.append('DATE')
            value.append(date)
            df = pd.DataFrame(value, index=index)
            print('df ------------------------------\n{}\n-------------------------------\n'.format(df))
            print('{} data have reshaped.'.format(len(data)))
            print('[evaluate] process has done.')

            return df


    def output(self, df_, train_target=None):

        print('\n[output]')
        print('='*30)
        if train_target == None:
            train_target = self.train_target
        if os.path.isfile(self.path_excel) is False:
            excel = openpyxl.Workbook()
            excel.save(self.path_excel)
            excel.close()
            self.processor_init()
            for i in range(len(self.comb[0])):
                self.init_sheet(i)
        sheet_name = train_target
        df = pd.read_excel(self.path_excel, sheet_name, index_col=0)
        df = df.merge(df_, left_index=True, right_index=True)
        with pd.ExcelWriter(self.path_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=train_target)

        print('[output] process has done.')


    def select_comb(self, use_num=False, use_all=False):

        comb, data_num = self.comb[0], self.comb[1]
        comb_trained = self.get_trained_comb()
        for i in range(len(comb))[::-1]:
            if comb[i] in comb_trained:
                    comb.pop(i)
                    data_num.pop(i)

        self.load_comb()
        for i in range(len(self.comb[0]))[::-1]:
            if data_num[i] != min(data_num):
                comb.pop(i)
                data_num.pop(i)

        if use_num is True:
            if use_all is True:
                return (comb, data_num)
            else:
                return (comb[0], data_num[0])

        if use_all is True:
            return comb
        else:
            return comb[0]


    # ホスト名のある行を取得 ('-' はホストなし)
    def get_trained_comb(self):

        sheet_name = 'process'
        df = pd.read_excel(self.path_excel, sheet_name, index_col=0)
        li = df.reset_index().values.tolist()
        _ = []
        for item in li:
            if item[1] != '-':
                _.append(item[0])
        return _


    def manage_host(self, mode='show', opt=None):

        df = pd.read_excel(self.path_excel, sheet_name='process', index_col=0)
        hostname = socket.gethostname()
        if mode == 'show':
            print('df---------------------\n{}\n-----------------------'.format(df))
        elif mode == 'regist':
            df.at[self.train_target, 'HOST'] = hostname
            with pd.ExcelWriter(self.path_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='process')
        elif mode == 'reset':
            if opt == 'all':
                df = pd.DataFrame(index=self.comb[0])
                df['HOST'] = '-'
            else:
                df.loc[df[df['HOST'] == hostname].index, 'HOST'] = '-'
            with pd.ExcelWriter(self.path_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='process')
        else:
            print('manage_host >> [{}] mode is undefined!'.format(mode))


    def print_data_num(self):

        self.load_comb()
        print('* combination : number')
        for i in range(len(self.comb[0])):
            print('  - [{}]  : {}'.format(self.comb[0][i], self.comb[1][i]))
        print('')


    def reset_excel(self):

        for i in range(len(self.comb[0])):
            self.init_sheet(i, replace=True)


    def write_log(self, file_name, message):
        with open(file_name, 'a') as f:
            print(message, file=f)


def main():

    os.system('cls')
    path = os.path.dirname(__file__) + '\\'
    data = node_comb(path)
    print('HOST: {}'.format(socket.gethostname()))
    print('path: {}'.format(path))

#   prep, tr, pred, eva, task='distractor', process_num=1
    data.start_process(process_num=1, train_target='RRRRRR')


if __name__ == '__main__':
    main()
